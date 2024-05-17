from datetime import datetime
import json
import os
import sys
import openpyxl
from constant import CategoryEnum
from util import get_previous_quarter, get_previous_two_quarter, upload_to_minio, call_send_email_api
#import excel2img
import excel
import news_excel

'''
lotes_report_text = {
"title": "Lotes 23年營運總結及24年展望-財務分析      20231228 rev.01",
"summary": """Summary: 
1. 營收: 
23年度營收 24,638MNTD, 較22年YOY-9%, 24年營收：28,934MNTD,較23年預估YOY+15%（預估), 25年營收較23年YOY+ 37%,
23~24年CAGR成長+8.4%。
2. 獲利：
2-1. 23年毛利率46.2%，較22年 +2.1%, 24年毛利率47%，較23年+0.8%   (預估)。（注：上漲原因為產品組合有利/營業費用維穩)
2-2.23年凈利率：25%，較22年+2%；24年凈利率：24.6%，較23年基本持平(預估)
3. By segment營收占比：
3-1.By產業看,24年伺服器產業擴大,  PC產業縮減,伺服器產業增長彌補PC產業的縮減, 伺服器為 2024 年主要成長動能
4. 24年營收獲利展望：
4-1. 24年展望 正向 。公司業務聚焦桌機、伺服器CPU、DDR等插槽與PCIe連接器。
4-2. PCIE連接器的市佔率由目前之10%提升至2024年15-20%,推升PCIe營收貢獻。
4-3. 伺服器營收成長顯著, 24年Q4達到全年最高峰（主因 CPU廠Intel(美)與AMD(美)預期4Q23新平台(Sapphire Rapid與Genoa)將放量。預估2024年伺服器營收年增40%-50%,推動提升2024年毛利率、營利率。
4-4.伺服器與桌機新平台升級推升插槽均價，毛利率仍將提升，2024年達47%, 較23年基本持平。
5. 未來成長動能
5-1. Intel新伺服器平台Birch Stream(可能於2025年推出)針腳數將增加至7,529(較Eagle Stream增加60%)，此將進一步推升CPU插槽均價攀升。
5-2. 隨PCIe Gen 5升級趨勢，嘉澤產品線除既有之連接器，將延伸至高速連接線產品，以連接GPU與CPU插槽 (AI伺服器)，並取代PCB。
5-3.材料價格上漲帶來一定的挑戰""",
"profit_summary": """獲利分析：
1. 23年營收獲利總結：
1-1. 23H1年營收下降(伺服器、桌機、筆電需求平淡),但毛利率穩定, 水準維持在45%的毛利率，23H2 營收上升15%,毛利上升2%
1-2. 23Q3 精利較H1提升7.5%, 營業成本貢獻3%,(產品組合有利)
1-3.23年營收下降,毛利穩定的主因：
①主要是新伺服器中央處理器（CPU）平台貢獻比重提高 
②新台幣貶值有利
③H2受惠桌機與伺服器急單，Q4新機種出貨帶動營收持續增長
""",
"by_segment_summary" : """By segment獲利占比：
1-1. By產業看,獲利佔比近3年微浮, 整體未來走向沒有大的偏移, 主要在PC及伺服器業務。
1-2. 23年By產業獲利佔比: PC 44%，伺服器26%，其他(工業+策略客戶) 15%，嘉基 10% ,汽車5%。
1-3. 24年伺服器獲利佔比增加4%，PC產業獲利佔比減少4%,其他產業維持23年水準,伺服器為 2024 年主要成長動能
1.4.23年由PC換代帶來的單價增長以外，預測市場佔有大幅成長，需PM/sales做進一步調查分析
"""
}
    
argosy_report_text = {
"title": "Argosy 23年營運總結及24年展望- 財報分析      20231228 rev.01",
"summary": """競爭對手趨勢分析- Argosy
Summary:
1.營收：
  a.23年度營收 3,006Mntd，24年預估營收 3,432Mntd，較23年上漲14%。
2.獲利：
  a.23年毛利額(率)1,387Mntd(46.1%)， 預計24年毛利額(率)1,625Mntd(47.3%)，較23Y上漲17%(1.2%)。
  b.23年凈利額(率)   741Mntd(24.7%)， 預計24年凈利額(率)   883Mntd(25.7%)，較23Y上漲19%(1.0%)。
3.By 系列營收佔比：
22年Top3：DDR 44%(SO-DIMM 37%/Long DIMM   7% )，M.2 28%，Type-C 11%；
23年Top3：DDR 50%(SO-DIMM 38%/Long DIMM 12% )，M.2 22%，Type-C 10%；
24年Top3：DDR 55%(SO-DIMM 34%/Long DIMM 21% )，M.2 19%，Type-C 9%；
DDR比重逐年增加，新世代DDR5的切換率23Y已達25%,預計24Y達到45%-50%，規格升級正確確保市場競爭力的延續(較高的ASP以及毛利率)；
4.優勢：
  a.DDR5 & 微型沖壓件(Micro stamping)等高毛利產品營收佔比持續提升；
  b.NB客戶于23Y H2 上修需求，新產品規格升級正確；
5.風險：
  出貨不如預期，ASP不如預期(新世代切換率較預期還有差距)
"""
}
    
tel_report_text = {
"title": "TEL 23年營運總結及24年展望- 財報分析      20231228 rev.01",
"summary": """競爭對手趨勢分析- TEL
Summary:
1.營收：
  a.23年度營收 16,034Musd，24年預估營收 16,903Musd，較23年上漲5.4%。
2.獲利：
  a.23年毛利額(率)5,055Musd(31.5%)， 預計24年毛利額(率)5,609Musd(33.2%)，較23Y上漲11%(1.7%)。
  b.23年凈利額(率)1,910Musd(11.9%)， 預計24年凈利額(率)2,398Musd(14.2%)，較23Y上漲25%(2.3%)。
3.By Segment 營收佔比：
22Y：Transportation   9,219Musd(56.6%),  Industrial 4,490Musd(27.6%),  Communications 2,572Musd(15.8%)；
23Y：Transportation   9,588Musd(59.8%),  Industrial 4,551Musd(28.4%),  Communications 1,895Musd(11.8%)；
24Y：Transportation 10,354Musd(61.3%),  Industrial 4,7290Musd(28.0%),Communications 1,820Musd(10.8%)；
運輸業比重逐年增加，營收由9,219M->10,354M,持續在電動汽車連接器,傳感器等領域佔據領先地位；
4.優勢：
  a.運輸業營收為9,219Musd穩步成長(同比成長4%)；
  b.通信業因AI注入新活力,23Y Q4較Q3成長4%；
  c.工業因介入手術的增加有利于醫療設備市場的需求增長 & 北美地區推動能源的有機增長,可再生能源設備應用持續成長；
5.風險：
  工業設備市場季節性疲軟,通訊類雖有回暖,但仍不如預期(各段供應鏈持續消減庫存)
"""
}    
    
aph_report_text = {
"title": "APH 23年營運總結及24年展望- 財報分析      20231228 rev.01",
"summary": """Summary:
營收：
1.23年度營收12,277MUSD, 较22年YOY-3%，24年預估營收 12,858MUSD，较23年YOY+3.5%，23~24年CAGR成長+2%。
獲利：
2-1. 23年毛利率32.3%，較22年YOY +0.4%, 24年毛利率32.7%，較23年+0.4%   (預估)。
2-2.23年凈利率：14.7%，較22年持平；24年凈利率：15.5%，較23年YOY +0.8% (預估)。
By segment：
23年占營收比Top3：工業類25%; 汽車23%;   網路通訊19%。
23年产业向工业及汽车产 方向转移, 工业类占营收比上涨2%，汽车占营收比上涨3%，网络通信占营收比下降3%; 24年營收類別基本與23年持平；
23年利潤貢獻點：
內部：1.產品組合(高毛利), 2.成本控制  3.較小程度的定價優勢; 4.庫存管控：庫存下降1.07億美金,庫存天數環比下降6.1天。
周邊：1.關閉RFS ,預計在23H2貢獻移動網路終端市場3000萬美金；2.關閉EBY Electro, 預計在工業終端市場貢獻1500萬美金的收入。
           3.Amphenol回購安費諾1.54億美金的股票,並支付利息1.25億美金。
24年營收獲利展望：
1.工業/移動網絡領域正在消化庫存,24年主要獲利貢獻主要看好汽車產業及IT通訊""",
"profit_summary": """獲利分析：
1.雖然終端市場不樂觀導致的營收下滑3%,但毛利及精利率穩定且小幅增長，
特別是在23Q3終端市場總體方向回穩,訂單連續三個季度擴大。
2.23Q4汽車及IT通訊終端市場保持平穩,對獲利貢獻正向。
""",
"by_segment_summary" : """By segment獲利占比：
1.By產業看,獲利佔比近3年, 汽車產業增長3%；軍工產業增長3%, 在獲利較好的產業內保持增長態勢。
2.By segment占營收 :汽車產業產23%, 工業25%,網絡通訊19%, 軍工11%,
24年同水平
"""
}   
'''

lotes_report_text = {
"title": " ",
"summary": """  """,
"profit_summary": """ """,
"by_segment_summary" : """ """
}
    
argosy_report_text = {
"title": " ",
"summary": """  """,
"profit_summary": """ """,
"by_segment_summary" : """ """
}
    
tel_report_text = {
"title": " ",
"summary": """  """,
"profit_summary": """ """,
"by_segment_summary" : """ """
}    
    
aph_report_text = {
"title": " ",
"summary": """  """,
"profit_summary": """ """,
"by_segment_summary" : """ """
}   

def init_text(fiscal_year, quarter):
  today = datetime.now().strftime("%Y%m%d")
  #lotes_report_text['title'] = f'Lotes {fiscal_year}年營運總結及{fiscal_year + 1}年展望- 財報分析      {today} rev.01'

def main(fiscal_year, quarter):
    init_text(fiscal_year, quarter)

    # 寄信時的參數
    mail_parameters = {
      "year": fiscal_year,
      "quarter": f"Q{quarter}"
    }
    # 寄信時的附檔
    attachments = []
    # 信件內容的圖片
    embedded_images = []

    # Load the Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Lotes sheet
    sheet.title = "Lotes"
    temporary_file = excel.create_lotes_style_financial_statements(sheet, "3533", fiscal_year, quarter, lotes_report_text, CategoryEnum.TW)
    for f in temporary_file:
        object_name = upload_to_minio(f)
        image_info = {"company": "Lotes", "file-name": object_name}
        embedded_images.append(image_info)

    # 3217 sheet
    sheet_3217 = wb.create_sheet("Argosy")
    #excel.create_argosy_style_financial_statements(sheet_3217, "3217", fiscal_year, quarter, argosy_report_text, CategoryEnum.TW)
    temporary_file = excel.create_lotes_style_financial_statements(sheet_3217, "3217", fiscal_year, quarter, argosy_report_text, CategoryEnum.TW)
    for f in temporary_file:
        object_name = upload_to_minio(f)
        image_info = {"company": "Argosy", "file-name": object_name}
        embedded_images.append(image_info)

    # TEL sheet
    sheet_tel = wb.create_sheet("TEL")
    #excel.create_argosy_style_financial_statements(sheet_tel, "tel", fiscal_year, quarter, tel_report_text, CategoryEnum.FOREIGN)
    temporary_file = excel.create_lotes_style_financial_statements(sheet_tel, "1385157", fiscal_year, quarter, tel_report_text, CategoryEnum.FOREIGN)
    for f in temporary_file:
      object_name = upload_to_minio(f)
      image_info = {"company": "TEL", "file-name": object_name}
      embedded_images.append(image_info)
    
    # APH
    sheet_aph = wb.create_sheet("APH")
    temporary_file = excel.create_lotes_style_financial_statements(sheet_aph, "820313", fiscal_year, quarter, aph_report_text, CategoryEnum.FOREIGN)    
    for f in temporary_file:
      object_name = upload_to_minio(f)
      image_info = {"company": "APH", "file-name": object_name}
      embedded_images.append(image_info)    

    # Save the Excel file
    wb.save('output/financial_statements.xlsx')

    object_name = upload_to_minio('output/financial_statements.xlsx')
    attachment_info = {"file-name": object_name}
    attachments.append(attachment_info)

    #excel2img.export_img("output/financial_statements.xlsx", "output/Lotes.png", "Lotes", None)
    #excel2img.export_img("output/financial_statements.xlsx", "output/Argosy.png", "Argosy", None)
    #excel2img.export_img("output/financial_statements.xlsx", "output/TEL.png", "TEL", None)
    #excel2img.export_img("output/financial_statements.xlsx", "output/APH.png", "APH", None)

    #upload_to_minio('output/Lotes.png')
    #upload_to_minio('output/Argosy.png')
    #upload_to_minio('output/TEL.png')
    #upload_to_minio('output/APH.png')    

    # create news excel file
    wb1 = openpyxl.Workbook()
    sheet = wb1.active
    
    # Lotes sheet
    sheet.title = "Lotes News"
    news_excel.create_news_sheet(sheet, '3533', fiscal_year, quarter, CategoryEnum.TW)

    # 3217 sheet
    sheet_3217 = wb1.create_sheet("Argosy News")
    news_excel.create_news_sheet(sheet_3217, '3217', fiscal_year, quarter, CategoryEnum.TW)

    # TEL sheet
    sheet_tel = wb1.create_sheet("TEL News")
    news_excel.create_news_sheet(sheet_tel, 'TEL', fiscal_year, quarter, CategoryEnum.FOREIGN)

    # APH
    sheet_aph = wb1.create_sheet("APH News")
    news_excel.create_news_sheet(sheet_aph, 'APH', fiscal_year, quarter, CategoryEnum.FOREIGN)  

    # Save the Excel file
    wb1.save('output/news.xlsx')

    object_name = upload_to_minio('output/news.xlsx')
    attachment_info = {"file-name": object_name}
    attachments.append(attachment_info)

    mail_parameters["embedded-images"] = embedded_images
    mail_parameters["attachments"] = attachments

    json_data = json.dumps(mail_parameters, indent=4, ensure_ascii=False)
    print(json_data)

    # Send email
    call_send_email_api(json_data)

if __name__ == "__main__":
  fiscal_year = None
  quarter = None

  if len(sys.argv) == 3:
    (fiscal_year, quarter) = sys.argv[1:]  

  if fiscal_year is None or quarter is None:
    current_date = datetime.now()
    fiscal_year = current_date.year
    quarter = (current_date.month - 1) // 3 + 1
  else :
    fiscal_year = int(fiscal_year)
    quarter = int(quarter)
  
  print(f"開始產生報表 {datetime.today()}")
  print(f"現在年度: {fiscal_year}  現在季度: {quarter}")
  (target_year, target_quarter) = get_previous_quarter(fiscal_year, quarter)
  print(f"財報年度: {target_year}  財報季度: {target_quarter}") 
  main(target_year, target_quarter)  
