from constant import CategoryEnum
import data_service
from models import CrawlerNews
from openpyxl.styles import Alignment

def __get_quarter_start_date(fiscal_year, quarter):
    if quarter == 1:
        return str(fiscal_year) + '-01-01'
    elif quarter == 2:
        return str(fiscal_year) + '-04-01'
    elif quarter == 3:
        return str(fiscal_year) + '-07-01'
    elif quarter == 4:
        return str(fiscal_year) + '-10-01'

def __get_quarter_end_date(fiscal_year, quarter):
    if quarter == 1:
        return str(fiscal_year) + '-03-31'
    elif quarter == 2:
        return str(fiscal_year) + '-06-30'
    elif quarter == 3:
        return str(fiscal_year) + '-09-30'
    elif quarter == 4:
        return str(fiscal_year) + '-12-31'
    
def __get_start_datetime(date):
    return date + ' 00:00:00'

def __get_end_datetime(date):
    return date + ' 23:59:59'


def __draw_news_sheet_tw(sheet, data: CrawlerNews):
    sheet['A1'] = "Title"
    sheet['B1'] = "Content"
    #sheet['C1'] = "Summary"
    sheet['C1'] = "Link"
    sheet['D1'] = "Publish Time"

    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 60

    row_index = 2

    for news in data:
        if news.title is not None:
            sheet.cell(row_index, 1).value = news.title
        if news.cleared_content is not None:
            sheet.cell(row_index, 2).value = news.cleared_content
        #if news.summary is not None:
        #    sheet.cell(row_index, 3).value = news.summary
        if news.link is not None:
            sheet.cell(row_index, 3).value = news.link
        if news.publish_on is not None:
            sheet.cell(row_index, 4).value = news.publish_on.strftime("%Y-%m-%d %H:%M:%S")

        # 設定自動換行
        for i in range(1, 5):
            sheet.cell(row_index, i).alignment = Alignment(wrap_text=True)

        row_index += 1            

def __draw_news_sheet_foreign(sheet, data: CrawlerNews):
    sheet['A1'] = "Title"
    sheet['B1'] = "Original"
    #sheet['C1'] = "Translation"
    #sheet['D1'] = "Summary"
    sheet['C1'] = "Link"
    sheet['D1'] = "Publish Time"

    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 60

    row_index = 2

    for news in data:
        if news.title is not None:
            sheet.cell(row_index, 1).value = news.title
        if news.cleared_content is not None:
            sheet.cell(row_index, 2).value = news.cleared_content
        #if news.translate_content is not None:
        #    sheet.cell(row_index, 3).value = news.translate_content
        #if news.summary is not None:
        #    sheet.cell(row_index, 4).value = news.summary
        if news.link is not None:
            sheet.cell(row_index, 3).value = news.link
        if news.publish_on is not None:
            sheet.cell(row_index, 4).value = news.publish_on.strftime("%Y-%m-%d %H:%M:%S")

        # 設定自動換行
        for i in range(1, 5):
            sheet.cell(row_index, i).alignment = Alignment(wrap_text=True)   

        row_index += 1    

def create_news_sheet(sheet, company_code, fiscal_year, quarter, category: CategoryEnum):
    quarter_start_datetime = __get_start_datetime(__get_quarter_start_date(fiscal_year, quarter))
    quarter_end_datetime = __get_end_datetime(__get_quarter_end_date(fiscal_year, quarter))

    news = data_service.get_news_by_stock_code_and_publish_on_between(company_code, quarter_start_datetime, quarter_end_datetime)

    if category == CategoryEnum.TW:
        __draw_news_sheet_tw(sheet, news)
    else:
        __draw_news_sheet_foreign(sheet, news)  

