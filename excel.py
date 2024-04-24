import math

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from io import BytesIO
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font
from sqlalchemy import text
from openpyxl.styles import PatternFill
from matplotlib.font_manager import FontProperties
from constant import CategoryEnum
import data_service
import platform
from PIL import Image as PILImage
from util import get_previous_quarter
import setting 


# 中文字型檔，圖表的Legend需要
os_name = platform.system()
if os_name == 'Windows':
    print("当前操作系统为 Windows")
    font = FontProperties(fname=r"C:\Windows\Fonts\mingliu.ttc", size=14)
else :
    print("当前操作系统为 Linux")
    font = FontProperties(fname=r"/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", size=14)

# excel邊框格式
border_style_tb = Border(top=Side(style='thin'),bottom=Side(style='thin'),)
border_style_ltb = Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),)
border_style_rtb = Border(top=Side(style='thin'),bottom=Side(style='thin'),right=Side(style='thin'),)
border_style_r = Border(right=Side(style='thin'),)
border_style_l = Border(left=Side(style='thin'),)
border_style_lt = Border(top=Side(style='thin'),left=Side(style='thin'),)
border_style_rt = Border(top=Side(style='thin'),right=Side(style='thin'),)
border_style_lb = Border(left=Side(style='thin'),bottom=Side(style='thin'),)
border_style_rb = Border(right=Side(style='thin'),bottom=Side(style='thin'),)
border_style_t = Border(top=Side(style='thin'),)
border_style_b = Border(bottom=Side(style='thin'),)
border_style_all = Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),)
   

def __is_multiple_row(sheet, cell_range):
    rows = sheet[cell_range]
    if len(rows) > 1:
        return True
    return False

def __draw_first_row_border(row):
    is_first_cell = True
    for cell in row:
        if is_first_cell:
            cell.border = border_style_lt
            is_first_cell = False
        else:
            cell.border = border_style_t
    cell.border = border_style_rt


def __draw_last_row_border(row):
    is_first_cell = True
    for cell in row:
        if is_first_cell:
            cell.border = border_style_lb
            is_first_cell = False
        else:
            cell.border = border_style_b
    cell.border = border_style_rb

def __draw_middle_row_border(row):
    is_first_cell = True
    for cell in row:
        if is_first_cell:
            cell.border = border_style_l
            is_first_cell = False
    cell.border = border_style_r

def __draw_single_row_border(row):
    is_first_cell = True
    for cell in row:
        if is_first_cell:
            cell.border = border_style_ltb
            is_first_cell = False
        else:
            cell.border = border_style_tb
    cell.border = border_style_rtb

def __draw_border(sheet, cell_range):
    if __is_multiple_row(sheet, cell_range):
        is_first_row = True
        for row in sheet[cell_range]:
            if is_first_row:
                __draw_first_row_border(row)
                is_first_row = False
            else:
                __draw_middle_row_border(row)
        __draw_last_row_border(row)

    else:
        for row in sheet[cell_range]:
            __draw_single_row_border(row)
        

def __set_cell_height(sheet, cell):
    max_height = 0
    over_length = 0
    length_boundary = 80
    if cell.value:
        lines = str(cell.value).split('\n')
        height = max(len(lines) + 1, 1) * 14  # Adjust this value as needed
        max_height = max(max_height, height)

        for line in lines:
            if len(line) > length_boundary:
                over_length += 1         
    if max_height > 0:
        sheet.row_dimensions[cell.row].height = max_height + over_length
        cell.alignment = Alignment(wrap_text=True, vertical='top')

def __set_cell_blod(cell):
    cell.font = Font(bold=True)

def __set_cell_alignment_center(cell):
    cell.alignment = Alignment(horizontal='center')

def __get_first_cell(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            return cell

def __get_formated_number(number):
    """
    取得格式化的數字，用逗號作為千位分隔符.
    
    Parameters:
    number (int or float): The number to be formatted.
    
    Returns:
    str: The formatted number with commas as thousand separators.
    """     
    return "{:,.0f}".format(round(number))
    #return "{:,}".format(number)

def __get_percentage_number(number):
    return "{:.1%}".format(number)

# 在excel中填入財報科目
def __draw_financial_statements_items_name(sheet, start_cell, category):
    row_index = start_cell.row
    column_index = start_cell.column
    sheet.column_dimensions['C'].auto_size = True
    items = setting.get_financial_statement_items_dict(category)
    color_format = setting.get_financial_statement_items_color_format_dict(category)
    sheet.cell(row=row_index, column=column_index).value = "科目"

    for (k, v) in items.items():
        row_index += 1
        sheet.cell(row=row_index, column=column_index).value = v
        sheet.cell(row=row_index, column=column_index).border = border_style_all

    # style
    row_index = start_cell.row
    sheet.cell(row=row_index, column=column_index).border = border_style_all
    sheet.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=row_index, column=column_index).fill = PatternFill(fill_type='solid', start_color=color_format["title"])

    for k in items.keys():
        row_index += 1
        sheet.cell(row=row_index, column=column_index).border = border_style_all
        sheet.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
        if k in color_format:
            sheet.cell(row=row_index, column=column_index).fill = PatternFill(fill_type='solid', start_color=color_format[k])
        

# 填入季度財報
def __draw_financial_statements_items(sheet, start_cell, comany_code, fiscal_year, quarter, category: CategoryEnum):    
    data_dict = None

    # get value
    result = data_service.get_financial_statements_quarter_values(comany_code, fiscal_year, quarter, category)
    
    # draw cell
    __draw_financial_statements_cells(sheet, start_cell, str(fiscal_year) + "Q" + str(quarter), result, category)

# 在excel中填入季度財報欄位
def __draw_financial_statements_cells(sheet, start_cell, title, data_dict, category):
    """
    Excel中填入季度財報數據的Cell
    """    
    row_index = start_cell.row
    column_index = start_cell.column
    #sheet.cell(row=row_index, column=column_index).value = str(fiscal_year) + "Q" + str(quarter)
    sheet.cell(row=row_index, column=column_index).value = title

    if category == CategoryEnum.TW:
        financial_statement_items = setting.get_financial_statement_items_dict(category)
        financial_statement_items_percentage_format = setting.get_financial_statement_items_percentage_format_dict(category)
        financial_statement_items_color_format = setting.get_financial_statement_items_color_format_dict(category)    
    else:
        financial_statement_items = setting.get_financial_statement_items_dict(category)
        financial_statement_items_percentage_format = setting.get_financial_statement_items_percentage_format_dict(category)
        financial_statement_items_color_format = setting.get_financial_statement_items_color_format_dict(category)

    # value
    if data_dict is not None:
        for (k, v) in financial_statement_items.items():
            row_index += 1
            if k in data_dict and data_dict[k] is not None:
                if k in financial_statement_items_percentage_format:
                    sheet.cell(row=row_index, column=column_index).value = __get_percentage_number(data_dict[k])
                else:
                    # TW財報數字單位與資料庫數字不同，需除以1000
                    if category == CategoryEnum.TW:
                        sheet.cell(row=row_index, column=column_index).value = __get_formated_number(data_dict[k] / 1000)
                    else:
                        sheet.cell(row=row_index, column=column_index).value = __get_formated_number(data_dict[k])
    
    # style
    row_index = start_cell.row
    sheet.cell(row=row_index, column=column_index).border = border_style_all
    sheet.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=row_index, column=column_index).fill = PatternFill(fill_type='solid', start_color=financial_statement_items_color_format["title"])

    for k in financial_statement_items.keys():
        row_index += 1
        sheet.cell(row=row_index, column=column_index).border = border_style_all
        sheet.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
        if k in financial_statement_items_color_format:
            sheet.cell(row=row_index, column=column_index).fill = PatternFill(fill_type='solid', start_color=financial_statement_items_color_format[k])

# 填入年度財報
def __draw_financial_statements_summary(sheet, start_cell, comany_code, fiscal_year, category):
    # get value
    data_dict = data_service.get_financial_statements_year_values(comany_code, fiscal_year, category)
    
    # draw cell
    __draw_financial_statements_cells(sheet, start_cell, str(fiscal_year), data_dict, category)

# Todo 將起始的cell傳入    
# 填入國內財報(季報+年報)
def __draw_financial_statements_tw(sheet, comany_code, input_fiscal_year, input_quarter):
    start_cell = sheet['C7']
    row_index = start_cell.row
    column_index = start_cell.column

#    __draw_financial_statements_items_name(sheet, sheet.cell(row=row_index, column=column_index), financial_statement_tw_items)
    __draw_financial_statements_items_name(sheet, sheet.cell(row=row_index, column=column_index), CategoryEnum.TW)

    # 畫季報
    quarter_array = __get_report_quarter(input_fiscal_year, input_quarter)
    for (year, quarter) in quarter_array:
        column_index += 1
        __draw_financial_statements_items(sheet, sheet.cell(row=row_index, column=column_index), comany_code, year, quarter, CategoryEnum.TW)

    # 畫年報
    year_array = __get_report_year(input_fiscal_year)
    for year in year_array:
        column_index += 1
        __draw_financial_statements_summary(sheet, sheet.cell(row=row_index, column=column_index), comany_code, year, CategoryEnum.TW)

# Todo 將起始的cell傳入  
# 填入國外財報(季報+年報) 
def __draw_financial_statements_foreign(sheet, comany_code, input_fiscal_year, input_quarter):
    start_cell = sheet['C7']
    row_index = start_cell.row
    column_index = start_cell.column

    #__draw_financial_statements_items_name(sheet, sheet['C7'], financial_statement_foreign_items)
    __draw_financial_statements_items_name(sheet, sheet['C7'], CategoryEnum.FOREIGN)

    # 畫季報
    quarter_array = __get_report_quarter(input_fiscal_year, input_quarter)
    for (year, quarter) in quarter_array:
        column_index += 1
        __draw_financial_statements_items(sheet, sheet.cell(row=row_index, column=column_index), comany_code, year, quarter, CategoryEnum.FOREIGN)

    # 畫年報
    year_array = __get_report_year(input_fiscal_year)
    for year in year_array:
        column_index += 1
        __draw_financial_statements_summary(sheet, sheet.cell(row=row_index, column=column_index), comany_code, year, CategoryEnum.FOREIGN)    

# 依據內容長度調整excel欄寬
def __autofit_column(ws, column_letter):
    column_index = openpyxl.utils.column_index_from_string(column_letter)
    max_length = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

    # Set the column width based on the maximum length
    ws.column_dimensions[column_letter].width = max_length * 1.75


# 產生圖表
def __get_combo_chart(categories, bar_chart_data, line_chart_data_1, line_chart_data_2):
    if os_name == 'Windows':
        plt.rcParams['font.family'] = 'mingliu'
    else: 
        plt.rcParams['font.family'] = 'WenQuanYi Micro Hei'

    fig, ax1 = plt.subplots()

    # Plotting the bar chart
    bars = ax1.bar(categories, bar_chart_data, color='#C5E0B4', label='營業收入', align='center', width=0.5)

    max_height = max(bar_chart_data)

    # Display values on the bar chart
    for bar, value in zip(bars, bar_chart_data):
        #ax1.text(bar.get_x() + bar.get_width() / 2, value + 0.5, str(value), ha='center', va='bottom', color='black', fontsize=10)
        ax1.text(bar.get_x() + bar.get_width() / 2, max_height * 0.1, __get_formated_number(value), ha='center', va='bottom', color='black', fontsize=10)

    # Create a secondary y-axis for the line charts
    ax2 = ax1.twinx()

    # Plotting the line charts
    line1, = ax2.plot(categories, line_chart_data_1, color='#FFDDA4', marker='o', label='毛利', markersize=8)
    line2, = ax2.plot(categories, line_chart_data_2, color='#5B9BD5', marker='s', label='淨利', markersize=8)

    ax2.set_ylabel('', color='r', fontproperties=font)

    # Display values on the line charts
    for x, y in zip(categories, line_chart_data_1):
        ax2.text(x, y + 5, __get_percentage_number(y/100), ha='center', va='bottom', color='black', fontsize=10)
    for x, y in zip(categories, line_chart_data_2):
        ax2.text(x, y + 5, __get_percentage_number(y/100), ha='center', va='bottom', color='black', fontsize=10)

    # Combine legends from both axes
    lines = [bars, line1, line2]
    labels = [line.get_label() for line in lines]
    ax2.legend(lines, labels, loc='upper right')

    max_value = max(bar_chart_data)
    ax1.set_ylim(0, max_value * 3)
    ax2.set_ylim(-100, 100)  # Set the y-axis limits for the line charts

    ax1.set_yticklabels([])
    ax2.set_yticklabels([])

    # Show the plot
    plt.title('By季獲利率趨勢 NT$m')
    plt.tight_layout()  # Ensures that the plots do not overlap
    #plt.show()

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    return buffer

# 產生圖表
def __get_operating_combo_chart(categories, bar_chart_data_1, bar_chart_data_2, bar_chart_data_3, line_chart_data_1, line_chart_data_2):
    if os_name == 'Windows':
        plt.rcParams['font.family'] = 'mingliu'
    else:
        plt.rcParams['font.family'] = 'WenQuanYi Micro Hei'

    # Create figure and axes for bar chart
    fig, ax1 = plt.subplots()

    # 计算每个柱状图的宽度
    bar_width = 0.25
    index = np.arange(len(categories))

    # Plotting the bar chart
    bars1 = plt.bar(index - bar_width, bar_chart_data_1, width=bar_width, label='營收', color='#5B9BD5')
    bars2 = plt.bar(index, bar_chart_data_2, width=bar_width, label='毛利', color = '#ED7D31')
    bars3 = plt.bar(index + bar_width, bar_chart_data_3, width=bar_width, label='淨利', color = '#FFC000')

    # Display values on the bar chart
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval / 2, __get_formated_number(yval), ha='center', va='bottom')

    # Create a secondary y-axis for the line charts
    ax2 = ax1.twinx()

    # Plotting the line charts
    line1, = ax2.plot(categories, line_chart_data_1, color='#A5A5A5', marker='o', label='毛利', markersize=8)
    line2, = ax2.plot(categories, line_chart_data_2, color='#4472C4', marker='s', label='淨利', markersize=8)
    ax2.set_ylabel('', color='r', fontproperties=font)

    # Display values on the line charts
    for x, y in zip(categories, line_chart_data_1):
        ax2.text(x, y + 5, __get_percentage_number(y/100), ha='center', va='bottom', color='black', fontsize=10)
    for x, y in zip(categories, line_chart_data_2):
        ax2.text(x, y + 5, __get_percentage_number(y/100), ha='center', va='bottom', color='black', fontsize=10)

    # Combine legends from both axes
    lines = [bars1, bars2, bars3, line1, line2]
    labels = [line.get_label() for line in lines]
    ax2.legend(lines, labels, loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=5)

    max_value = max(bar_chart_data_1)
    ax1.set_ylim(0, max_value * 3)
    ax2.set_ylim(-100, 100)  # Set the y-axis limits for the line charts
    ax1.set_yticklabels([])
    ax2.set_yticklabels([])

    # Show the plot
    plt.title('營運指標趨勢')
    plt.tight_layout()  # Ensures that the plots do not overlap
    #plt.show()    

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    return buffer

def __extract_year_data(data):
    result = []
    for row in data:
        result.append(row[0]) 
    return result

def __get_quarter_profit_chart(company_code, quarter_array, category: CategoryEnum): 
    quarter_financial_statements = []
    for target_year, target_quarter in quarter_array:
        data = data_service.get_financial_statements_quarter_values(company_code, target_year, target_quarter, category)
        quarter_financial_statements.append(data)

    quarter_array_after_filter = [f"{data.get('fiscal_year')}Q{data.get('quarter')}" for data in quarter_financial_statements]    
    
    if category == CategoryEnum.TW:
        # TW財報單位與資料庫單位不同，需要除以1000
        operating_revenue_array = [ data.get('operating_revenue') / 1000 for data in quarter_financial_statements] 
    else:  
        operating_revenue_array = [ data.get('net_sales') for data in quarter_financial_statements] 
    
    gross_profit_margin_percentage = [data.get('gross_profit_margin') * 100 for data in quarter_financial_statements]
    net_profit_margin_percentage = [data.get('net_profit_margin') * 100 for data in quarter_financial_statements]
    
    return __get_combo_chart(quarter_array_after_filter, operating_revenue_array, gross_profit_margin_percentage, net_profit_margin_percentage)

def __get_year_profit_chart(company_code, start_fiscal_year, end_fiscal_year, category: CategoryEnum):
    year_financial_statements = []
    for i in range(start_fiscal_year, end_fiscal_year + 1):
        data = data_service.get_financial_statements_year_values(company_code, i, category)
        year_financial_statements.append(data)

    categories = [str(i) for i in range(start_fiscal_year, end_fiscal_year + 1)] 

    # TW財報單位與資料庫單位不同，需要除以1000
    if category == CategoryEnum.TW:        
        operating_revenue_array = [ i.get('operating_revenue') / 1000 for i in year_financial_statements]
    else:
        operating_revenue_array = [ i.get('net_sales') for i in year_financial_statements]

    gross_profit_margin_percentage = [i.get('gross_profit_margin') * 100 for i in year_financial_statements] 
    net_profit_margin_percentage = [i.get('net_profit_margin') * 100 for i in year_financial_statements]   
       
    return __get_combo_chart(categories, operating_revenue_array, gross_profit_margin_percentage, net_profit_margin_percentage)

def __get_year_operating_chart(company_code, start_fiscal_year, end_fiscal_year, category: CategoryEnum):
    if category == CategoryEnum.TW:
        operating_revenue = data_service.get_year_operating_revenue(company_code, start_fiscal_year, end_fiscal_year)
        gross_profit = data_service.get_year_gross_profit(company_code, start_fiscal_year, end_fiscal_year)
        net_income = data_service.get_year_net_income(company_code, start_fiscal_year, end_fiscal_year)
        gross_profit_margin = data_service.get_year_gross_profit_margin(company_code, start_fiscal_year, end_fiscal_year)
        net_profit_margin = data_service.get_year_net_profit_margin(company_code, start_fiscal_year, end_fiscal_year)
    else:
        operating_revenue = data_service.get_year_net_sales_foreign(company_code, start_fiscal_year, end_fiscal_year)
        gross_profit = data_service.get_year_gross_profit_foreign(company_code, start_fiscal_year, end_fiscal_year)
        net_income = data_service.get_year_net_income_foreign(company_code, start_fiscal_year, end_fiscal_year)
        gross_profit_margin = data_service.get_year_gross_profit_margin_foreign(company_code, start_fiscal_year, end_fiscal_year)
        net_profit_margin = data_service.get_year_net_profit_margin_foreign(company_code, start_fiscal_year, end_fiscal_year)

    categories = [str(i) for i in range(start_fiscal_year, end_fiscal_year + 1)]
    operating_revenue_array = __extract_year_data(operating_revenue)
    gross_profit_array = __extract_year_data(gross_profit)
    net_income_array = __extract_year_data(net_income)
    gross_profit_margin_array = __extract_year_data(gross_profit_margin)
    gross_profit_margin_percentage_array = [i * 100 for i in gross_profit_margin_array]
    net_profit_margin_array = __extract_year_data(net_profit_margin)
    net_profit_margin_percentage_arrary = [i * 100 for i in net_profit_margin_array]

    return __get_operating_combo_chart(categories, operating_revenue_array, gross_profit_array, net_income_array, gross_profit_margin_percentage_array, net_profit_margin_percentage_arrary)

def __get_report_year(input_fiscal_year):
    """
    取得要產生財報年報部分的年份
    Returns a list of integers representing the current year and the two previous and two subsequent years.
    """
    result = []
    for i in range(-1, 2):
        result.append(input_fiscal_year + i)
    
    return result

def __get_quarter(input_fiscal_year, input_quarter, diff):
    """
    Calculates the year and quarter based on a given difference.

    Args:
        diff (int): The difference in quarters from the current quarter.

    Returns:
        tuple: A tuple containing the year and quarter.

    Example:
        >>> __get_quarter(1)
        (2022, 4)
    """
    current_year = input_fiscal_year
    current_quarter = input_quarter
    

    diff_year = math.floor(abs(diff) / 4)
    remainder = abs(diff) % 4

    if diff < 0:
        remainder = - remainder
        diff_year = - diff_year

    year = current_year + diff_year
    quarter = current_quarter
    if current_quarter + remainder > 4:
        year += 1
    elif current_quarter + remainder < 1:
        year -= 1

    quarter = (current_quarter + diff) % 4 or 4 # current_quarter + diff

    return(year, quarter)

def __get_report_quarter(input_fiscal_year, input_quarter):
    quarters_array = []

    # 前8個季度(含本季度) + 後4個季度
    for i in range(-7, 5):        
        quarters_array.append(__get_quarter(input_fiscal_year, input_quarter, i))         

    return quarters_array

def __get_chart_year(input_fiscal_year):
    """
    取得要產生圖表的年份
    Returns a list of integers representing the current year and the two previous and two subsequent years.
    """
    result = []

    for i in range(-3, 0):
        result.append(input_fiscal_year + i)
    
    return result
 
def __get_chart_quarter(input_fiscal_year, input_quarter):
    quarters_array = []

    # 前8個季度(含本季度)
    for i in range(-7, 1):        
        quarters_array.append(__get_quarter(input_fiscal_year, input_quarter, i))         

    return quarters_array

def create_lotes_style_financial_statements(sheet, company_code, input_fiscal_year, input_quarter, report_text, category: CategoryEnum):
    # 需要上傳的暫存檔
    temporary_file = []

    for(k, v) in setting.lotes_report_format.items():
        # 合併儲存格
        if not __is_multiple_row(sheet, v):
            sheet.merge_cells(v)
        # 劃格線
        __draw_border(sheet, v)
        # 填入文字
        if report_text.get(k):
            __get_first_cell(sheet, v).value = report_text[k]
            __set_cell_height(sheet, __get_first_cell(sheet, v))    
        # 處理標題格式
        if k == "title":
            __set_cell_blod(__get_first_cell(sheet, v))    
            __set_cell_alignment_center(__get_first_cell(sheet, v))
    
    # 填入財報excel資料
    if category == CategoryEnum.TW:
        __draw_financial_statements_tw(sheet, company_code, input_fiscal_year, input_quarter)
    else:
        __draw_financial_statements_foreign(sheet, company_code, input_fiscal_year, input_quarter)

    # 填入獲利分析圖表  
    # 年度分析圖    
    year_array = __get_chart_year(input_fiscal_year)    
    buffer = __get_year_profit_chart(company_code, year_array[0], year_array[-1], category)

    img = openpyxl.drawing.image.Image(buffer)
    img.width = 300
    img.height = 300
    chart_location = __get_first_cell(sheet, setting.lotes_report_format['profit_summary_chart'])
    sheet.add_image(img, chart_location.coordinate)

    # 保存圖表
    file_name = f"output/{company_code}_year_chart.png"
    pil_image = PILImage.open(buffer)
    pil_image.save(file_name)
    temporary_file.append(file_name)

    # 季度分析圖  
    quarter_array = __get_chart_quarter(input_fiscal_year, input_quarter)
    buffer = __get_quarter_profit_chart(company_code, quarter_array, category)

    img = openpyxl.drawing.image.Image(buffer)
    img.width = 450
    img.height = 300
    row = chart_location.row
    col = chart_location.column + 5
    chart_location = sheet.cell(row, col)
    sheet.add_image(img, chart_location.coordinate)
    # 調整row高度
    # image.height的單位是pixel, cell.height的單位是point, 1 pixel = 0.75 point
    row_height = max(img.height * 0.75, sheet.row_dimensions[chart_location.row].height)
    sheet.row_dimensions[chart_location.row].height = row_height

    # 保存圖表
    file_name = f"output/{company_code}_quarter_chart.png"
    pil_image = PILImage.open(buffer)
    pil_image.save(file_name)
    temporary_file.append(file_name)

    # 調整財報"科目"欄位寬度
    __autofit_column(sheet, "C")

    return temporary_file

def create_argosy_style_financial_statements(sheet, company_code, input_fiscal_year, input_quarter, report_text, category: CategoryEnum):
    for(k, v) in setting.argosy_report_format.items():
        # 合併儲存格
        sheet.merge_cells(v)
        # 劃格線
        __draw_border(sheet, v)
        # 填入文字
        if report_text.get(k):
            __get_first_cell(sheet, v).value = report_text[k]
            __set_cell_height(sheet, __get_first_cell(sheet, v))    
        # 處理標題格式
        if k == "title":
            __set_cell_blod(__get_first_cell(sheet, v))    
            __set_cell_alignment_center(__get_first_cell(sheet, v))

    # 填入獲利分析圖表
    year_array = __get_chart_year(input_fiscal_year)
    buffer = __get_year_operating_chart(company_code, year_array[0], year_array[-1], category)
    img = openpyxl.drawing.image.Image(buffer)
    img.width = 600
    img.height = 450
    sheet.add_image(img, "K4") 




